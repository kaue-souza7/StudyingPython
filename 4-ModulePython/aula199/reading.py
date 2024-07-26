# openpyxl - ler e alterar dados de uma planilha 
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

# Carregando um arquivo Excel
workbook = load_workbook(WORKBOOK_PATH)
# worksheet: Worksheet = workbook.active # type: ignore

# Nome da Planilha
sheetname = 'MinhaPlanilha'

# Selecionando a planilha
worksheet: Worksheet = workbook[sheetname] # type: ignore

# row: tuple[Cell]
for row in worksheet.iter_rows(): 
    for cell in row:
        print(cell.value, end='\t')

        # if cell.value == 'Maria':
        #     worksheet.cell(cell.row, 2, '20')
    print()

worksheet['B3'].value = 20





workbook.save(WORKBOOK_PATH)


# RECUPERANDO DADOS DA TABELA E JOGANDO EM UMA VARIAVEL

students_info = []

for row in worksheet.iter_rows():
    new_student = []
    for col in row:
         new_student.append(col.value)
    students_info.append(new_student)

for s in students_info:
    print(s)